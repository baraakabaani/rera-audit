"""Pull structured figures out of customer trial balances and bank statements.

The output feeds :mod:`app.services.workbook_generator`.  It is intentionally
conservative - it only reports a figure when a row's account label maps cleanly
to a canonical bucket, and it always records the source so the auditor can trace
every populated cell back to a file.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

_NUM_RE = re.compile(r"-?\(?\d[\d,]*\.?\d*\)?")

# canonical bucket -> trigger keywords (matched against the account label)
BUCKETS: dict[str, list[str]] = {
    "income_service_charge": ["service charge income", "service charge", "sc income", "levy income"],
    "income_chilled_water": ["chilled water", "in-unit chilled", "district cooling"],
    "income_other": ["other income", "miscellaneous income", "interest income", "fine", "penalty income"],
    "exp_services": ["security", "cleaning", "manpower", "guarding", "lifeguard", "concierge"],
    "exp_maintenance": ["maintenance", "amc", "repair", "lift maintenance", "hvac", "mep", "pest"],
    "exp_community_improvement": ["community improvement", "landscaping", "beautification"],
    "exp_utility": ["dewa", "empower", "sewa", "utility", "electricity", "water charges", "chilled water expense"],
    "exp_management_fee": ["management fee", "mc fee", "supervision fee"],
    "exp_insurance": ["insurance"],
    "exp_master_community": ["master community", "nakheel", "emaar community", "master developer"],
    "exp_reserve_fund": ["reserve fund expense", "sinking fund expense", "replacement fund"],
    "exp_provision_ecl": ["expected credit loss", "provision for doubtful", "bad debt", "ecl"],
    "asset_sc_receivable": ["service charge receivable", "sc receivable", "trade receivable", "owners receivable"],
    "asset_provision_ecl": ["provision for expected credit loss", "allowance for ecl", "provision - ecl"],
    "asset_bank_general": ["general fund bank", "gf bank", "current account", "operating account", "bank - general"],
    "asset_bank_reserve": ["reserve fund bank", "rf bank", "bank - reserve", "regulated bank - reserve"],
    "asset_short_term_deposit": ["short term deposit", "fixed deposit", "term deposit", "wakala"],
    "asset_prepaid": ["prepaid", "prepayment", "advance to supplier"],
    "asset_deposits": ["security deposit", "refundable deposit"],
    "liab_accounts_payable": ["accounts payable", "trade payable", "sundry creditors", "supplier payable"],
    "liab_accrued": ["accrued", "accruals", "accrued expense"],
    "liab_sc_advance": ["received in advance", "advance service charge", "unearned", "deferred income"],
    "liab_security_deposit": ["security deposit payable", "tenant deposit", "refundable deposit payable"],
    "liab_retention": ["retention payable", "retention"],
}


@dataclass
class Figure:
    bucket: str
    label: str
    amount: float
    source_file: str
    source_ref: str = ""      # sheet!cell or row hint


@dataclass
class FinancialData:
    figures: list[Figure] = field(default_factory=list)
    bank_general_closing: float | None = None
    bank_reserve_closing: float | None = None
    suppliers: list[tuple[str, float]] = field(default_factory=list)   # (name, balance)
    notes: list[str] = field(default_factory=list)

    def total(self, bucket: str) -> float | None:
        vals = [f.amount for f in self.figures if f.bucket == bucket]
        return round(sum(vals), 2) if vals else None

    def as_dict(self) -> dict:
        return {
            "figures": [f.__dict__ for f in self.figures],
            "bank_general_closing": self.bank_general_closing,
            "bank_reserve_closing": self.bank_reserve_closing,
            "suppliers": self.suppliers,
            "notes": self.notes,
        }


def _to_float(raw) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s:
        return None
    m = _NUM_RE.search(s.replace(" ", ""))
    if not m:
        return None
    tok = m.group(0).replace(",", "")
    neg = tok.startswith("(") and tok.endswith(")")
    tok = tok.strip("()")
    try:
        val = float(tok)
    except ValueError:
        return None
    return -val if neg else val


def _bucket_for(label: str) -> str | None:
    low = label.lower()
    for bucket, kws in BUCKETS.items():
        if any(k in low for k in kws):
            return bucket
    return None


def _iter_rows(path: Path):
    """Yield (sheet, row_index, list[str|num]) for xlsx/xls/csv."""
    ext = path.suffix.lower()
    if ext == ".xlsx":
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        for sn in wb.sheetnames:
            for i, row in enumerate(wb[sn].iter_rows(values_only=True), start=1):
                yield sn, i, list(row)
        wb.close()
    else:
        import pandas as pd

        if ext == ".csv":
            frames = {"CSV": pd.read_csv(path, dtype=str, header=None, nrows=4000, on_bad_lines="skip")}
        else:
            frames = pd.read_excel(path, sheet_name=None, dtype=str, header=None, engine="xlrd")
        for sn, df in frames.items():
            for i, row in enumerate(df.values.tolist(), start=1):
                yield sn, i, row


def _looks_like_tb(rows: list[tuple[str, int, list]]) -> bool:
    head = " ".join(
        str(c).lower() for _, _, r in rows[:15] for c in r if c is not None
    )
    return ("trial balance" in head) or (
        ("debit" in head and "credit" in head) or ("balance" in head and "account" in head)
    )


def parse_financials(doc_paths: list[tuple[Path, str, str]]) -> FinancialData:
    """``doc_paths``: list of (path, filename, folder)."""
    data = FinancialData()

    for path, fname, folder in doc_paths:
        if path.suffix.lower() not in (".xlsx", ".xls", ".csv"):
            continue
        try:
            rows = list(_iter_rows(path))
        except Exception as exc:  # noqa: BLE001
            data.notes.append(f"{fname}: could not read ({exc})")
            continue

        name_l = f"{fname} {folder}".lower()
        is_bank = any(k in name_l for k in ("bank", "adcb", "enbd", "statement", "ledger"))
        is_tb = _looks_like_tb(rows) or "trial balance" in name_l

        if is_tb:
            _harvest_tb(rows, fname, data)
        if is_bank:
            _harvest_bank(rows, fname, name_l, data)
        if any(k in name_l for k in ("supplier", "payable", "creditor", "soa", "aging", "ageing")):
            _harvest_suppliers(rows, fname, data)

    return data


def _harvest_tb(rows, fname: str, data: FinancialData) -> None:
    found = 0
    for sn, i, row in rows:
        cells = [c for c in row if c is not None and str(c).strip() != ""]
        if len(cells) < 2:
            continue
        label = str(cells[0]).strip()
        if len(label) < 3 or _to_float(label) is not None:
            continue
        bucket = _bucket_for(label)
        if not bucket:
            continue
        nums = [v for v in (_to_float(c) for c in cells[1:]) if v is not None]
        if not nums:
            continue
        amount = max(nums, key=abs)
        data.figures.append(
            Figure(bucket=bucket, label=label, amount=abs(amount), source_file=fname, source_ref=f"{sn}!row{i}")
        )
        found += 1
    if found:
        data.notes.append(f"{fname}: mapped {found} trial-balance line(s).")


def _harvest_bank(rows, fname: str, name_l: str, data: FinancialData) -> None:
    is_reserve = any(k in name_l for k in ("reserve", " rf", "rf ", "sinking"))
    closing = None
    for _sn, _i, row in rows:
        joined = " ".join(str(c).lower() for c in row if c is not None)
        if any(k in joined for k in ("closing balance", "ending balance", "balance c/f", "balance carried")):
            nums = [v for v in (_to_float(c) for c in row) if v is not None]
            if nums:
                closing = nums[-1]
    if closing is None:
        # fall back to the last numeric value in the last populated row
        for _sn, _i, row in reversed(rows):
            nums = [v for v in (_to_float(c) for c in row) if v is not None]
            if nums:
                closing = nums[-1]
                break
    if closing is None:
        return
    if is_reserve:
        data.bank_reserve_closing = closing
        data.notes.append(f"{fname}: reserve-fund bank closing balance {closing:,.2f}.")
    else:
        data.bank_general_closing = closing
        data.notes.append(f"{fname}: general-fund bank closing balance {closing:,.2f}.")


def _harvest_suppliers(rows, fname: str, data: FinancialData) -> None:
    count = 0
    for _sn, _i, row in rows:
        cells = [c for c in row if c is not None and str(c).strip() != ""]
        if len(cells) < 2:
            continue
        name = str(cells[0]).strip()
        if len(name) < 3 or _to_float(name) is not None or name.lower() in {"total", "supplier name", "supplier"}:
            continue
        nums = [v for v in (_to_float(c) for c in cells[1:]) if v is not None]
        if not nums:
            continue
        data.suppliers.append((name, abs(max(nums, key=abs))))
        count += 1
    if count:
        data.notes.append(f"{fname}: extracted {count} supplier balance(s).")
