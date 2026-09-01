"""Produce a completed copy of the auditor's Requirements Checklist.

Takes the *original* requirements workbook (so every style, column width and the
``=TODAY()`` date formula are preserved) and fills in:

* **STATUS**  - from the match result (Provided / Provided (partial) / Pending / N/A)
* **REMARKS** - the mapped file name(s), or what is still outstanding
* the **Audit Period** banner - refreshed to the reporting period detected from
  the customer files

Saved as ``requirements_filled.xlsx``.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from ..schemas import EntityContext, MatchResult
from .requirements_parser import _find_table

_STATUS_WORD = {
    "Received": "Provided",
    "Partial": "Provided (partial)",
    "Pending": "Pending",
    "Not applicable": "Not applicable",
}
_FILL = {
    "Received": PatternFill("solid", fgColor="C6EFCE"),
    "Partial": PatternFill("solid", fgColor="FFEB9C"),
    "Pending": PatternFill("solid", fgColor="FFC7CE"),
    "Not applicable": PatternFill("solid", fgColor="E7E6E6"),
}
_FONT_RGB = {
    "Received": "FF006100",
    "Partial": "FF9C5700",
    "Pending": "FF9C0006",
    "Not applicable": "FF3F3F3F",
}


def _remark(row) -> str:
    if row.status == "Received":
        files = ", ".join(m.filename for m in row.matched_files)
        return f"Provided - {files}" if files else "Provided."
    if row.status == "Partial":
        files = ", ".join(m.filename for m in row.matched_files)
        base = f"Partially provided - {files}. " if files else ""
        return (base + (row.comment or "Auditor to confirm completeness.")).strip()
    if row.status == "Not applicable":
        return row.comment or "Not applicable to this JOP."
    return row.comment or "Not yet provided by the client."


def fill_requirements(
    src_path: str | Path,
    out_path: str | Path,
    match: MatchResult,
    context: EntityContext | None = None,
    sheet_name: str | None = None,
) -> dict:
    wb = openpyxl.load_workbook(str(src_path), data_only=False)
    sheet = sheet_name if sheet_name in wb.sheetnames else next(
        (s for s in wb.sheetnames if "requirement" in s.lower()), wb.sheetnames[0]
    )
    ws = wb[sheet]

    def anchor(row: int, col: int):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return ws.cell(row=rng.min_row, column=rng.min_col)
        return ws.cell(row=row, column=col)

    _hr, _cref, _creq, _cev, c_status, c_rem = _find_table(ws)

    by_row = {r.row: r for r in match.rows if r.row}
    by_ref = {r.ref: r for r in match.rows}
    written = 0
    counts: dict[str, int] = {}

    # match on the stored source-row first, fall back to fuzzy text match
    text_index = {
        re.sub(r"\W+", " ", (ws.cell(row=rr, column=_creq).value or "")).strip().lower(): rr
        for rr in range(_hr + 1, ws.max_row + 1)
    }

    used_rows: set[int] = set()
    for r in match.rows:
        rr = r.row if r.row in range(1, ws.max_row + 1) else None
        if rr is None:
            key = re.sub(r"\W+", " ", r.requirement).strip().lower()
            rr = text_index.get(key)
        if not rr or rr in used_rows:
            continue
        used_rows.add(rr)
        cell = anchor(rr, c_status)
        cell.value = _STATUS_WORD.get(r.status, r.status)
        if r.status in _FILL:
            cell.fill = _FILL[r.status]
            old = cell.font
            cell.font = Font(name=old.name, size=old.size, bold=old.bold, color=_FONT_RGB[r.status])
        anchor(rr, c_rem).value = _remark(r)[:400]
        written += 1
        counts[r.status] = counts.get(r.status, 0) + 1

    # refresh the "Audit Period : ..." banner
    period_updated = ""
    if context and context.period_end:
        rng = f"{context.period_start} to {context.period_end}"
        for row in ws.iter_rows(max_row=15):
            for cc in row:
                if type(cc).__name__ == "MergedCell":
                    continue
                if isinstance(cc.value, str) and "audit period" in cc.value.lower():
                    cc.value = re.sub(
                        r"(audit period\s*[:\-]?\s*).*", lambda m: m.group(1) + rng, cc.value, flags=re.I
                    )
                    period_updated = rng
                elif isinstance(cc.value, str) and re.search(r"6 months period ended", cc.value, re.I):
                    cc.value = re.sub(
                        r"(ended\s+).*", lambda m: m.group(1) + context.period_end.upper(),
                        cc.value, flags=re.I,
                    )

    out_path = Path(out_path)
    wb.save(str(out_path))
    wb.close()
    return {
        "filename": out_path.name,
        "rows_written": written,
        "status_counts": counts,
        "period_updated": period_updated,
    }
