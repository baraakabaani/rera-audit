"""LLM-driven workbook population (Groq / OpenAI-compatible / Anthropic).

Pipeline:

1. **Classify** each customer spreadsheet (trial balance, bank, invoice schedule,
   budget, collection report, GL extract ...).
2. **Extract** - one small-model call per financial file turns the raw rows into a
   structured JSON of accounting facts.
3. **Consolidate** the facts into a single FactBook.
4. **Map** - one main-model call per annexure sheet: given the sheet's exact
   layout + the FactBook, return the precise cell writes (values + Excel
   formulas) to populate every data cell.
5. **Reconcile** - one final call to fix the balance-sheet tie-out and any key
   blanks.

Everything degrades: a throttled or failed step just yields fewer writes.
"""
from __future__ import annotations

import json
import time
from pathlib import Path

from openpyxl.utils import get_column_letter

from .. import config
from ..schemas import EntityContext
from . import llm_client
from .document_extractor import full_table_text, full_text
from .financials import FinancialData

_SKIP_SHEETS = {
    "cover", "index", "audit report",
    "17. pbc checklist status", "18. requirements tracker", "16. notes summary",
}
_MAX_LAYOUT_ROWS = 70
_CALL_GAP_S = 1.2

_extract_cache: dict[tuple[str, int], dict] = {}


# --------------------------------------------------------------------------- #
# 1. classify
# --------------------------------------------------------------------------- #
def classify(filename: str, folder: str, ext: str) -> str:
    s = f"{filename} {folder}".lower()
    if ext not in (".xlsx", ".xls", ".csv"):
        return "document"
    if "trial balance" in s or s.strip().startswith("tb ") or " tb" in s:
        return "trial_balance"
    if any(k in s for k in ("bank", "adcb", "enbd", "ubl", "statement", "ledger")):
        return "bank"
    if any(k in s for k in ("invoiceschedule", "invoice schedule", "levy", "receivable", "outstanding")):
        return "invoice_schedule"
    if any(k in s for k in ("collection", "propertygroup", "mollak")):
        return "collection"
    if "budget" in s or "variance" in s:
        return "budget"
    if any(k in s for k in ("a1.", "a2.", "a4.", "a5.", "b1.", "b6.", "gl ", "general ledger", "journal")):
        return "gl_extract"
    return "other_sheet"


_FINANCIAL_KINDS = {"trial_balance", "bank", "invoice_schedule", "collection", "budget", "gl_extract"}


# --------------------------------------------------------------------------- #
# 2. extract
# --------------------------------------------------------------------------- #
_EXTRACT_SYSTEM = (
    "You extract the accounting numbers from a Dubai jointly-owned-property (JOP) "
    "service-charge accounting file for a 6-month interim RERA review. Read the "
    "rows and return the figures, resolving the hierarchy (section totals vs "
    "sub-accounts). All amounts as POSITIVE numbers (magnitudes). Use the CLOSING "
    "balance / period figure, not opening. Do not invent anything - omit what is "
    "not in the file."
)


def _extract_file(path: Path, kind: str, filename: str) -> dict:
    key = (str(path), path.stat().st_size if path.exists() else 0)
    if key in _extract_cache:
        return _extract_cache[key]
    table = full_table_text(path)
    if not table or table.startswith("(could not read"):
        _extract_cache[key] = {}
        return {}
    prov = llm_client.resolve()
    small = config.LLM_EXTRACT_MODEL or (
        "llama-3.1-8b-instant" if prov and prov[0] == "groq" else None
    )
    user = (
        f"FILE: {filename}\nLIKELY KIND: {kind}\n\nROWS:\n{table}\n\n"
        'Return JSON with only the keys you have data for:\n'
        '{"income":[{"label":"","amount":0}],'
        '"expenditure":[{"label":"","amount":0}],'
        '"assets":[{"label":"","amount":0}],'
        '"liabilities":[{"label":"","amount":0}],'
        '"fund_balances":{"general_fund_opening":0,"reserve_fund_opening":0},'
        '"bank":{"general_fund_closing":0,"reserve_fund_closing":0},'
        '"receivables_total":0,'
        '"suppliers":[{"name":"","balance":0}],'
        '"budget":[{"code":"","service":"","budget_annual":0,"budget_6m":0,"actual_6m":0}],'
        '"notes":[""]}'
    )
    try:
        data = llm_client.chat_json(_EXTRACT_SYSTEM, user, max_tokens=2400, model=small, retries=3)
    except Exception:
        data = {}
    _extract_cache[key] = data if isinstance(data, dict) else {}
    return _extract_cache[key]


# --------------------------------------------------------------------------- #
# 3. consolidate
# --------------------------------------------------------------------------- #
def _num(v):
    try:
        n = float(str(v).replace(",", "").replace("(", "-").replace(")", ""))
        return abs(n) if n else 0.0
    except (TypeError, ValueError):
        return None


def consolidate(objs: list[dict], fin: FinancialData, ctx: EntityContext) -> dict:
    fb: dict = {
        "period": f"{ctx.period_start} to {ctx.period_end}",
        "income": [], "expenditure": [], "assets": [], "liabilities": [],
        "fund_balances": {}, "bank": {}, "suppliers": [], "budget": [], "notes": [],
        "receivables_total": None,
    }
    seen_sup = set()
    for o in objs:
        if not isinstance(o, dict):
            continue
        for k in ("income", "expenditure", "assets", "liabilities"):
            for row in o.get(k, []) or []:
                amt = _num(row.get("amount"))
                lbl = str(row.get("label", "")).strip()
                if lbl and amt:
                    fb[k].append({"label": lbl[:60], "amount": round(amt, 2)})
        for row in o.get("budget", []) or []:
            fb["budget"].append({
                "code": str(row.get("code", ""))[:12], "service": str(row.get("service", ""))[:50],
                "budget_annual": _num(row.get("budget_annual")), "budget_6m": _num(row.get("budget_6m")),
                "actual_6m": _num(row.get("actual_6m")),
            })
        for s in o.get("suppliers", []) or []:
            name = str(s.get("name", "")).strip()[:48]
            bal = _num(s.get("balance"))
            if name and bal is not None and name.lower() not in seen_sup:
                seen_sup.add(name.lower())
                fb["suppliers"].append({"name": name, "balance": round(bal, 2)})
        for kk, vv in (o.get("fund_balances") or {}).items():
            n = _num(vv)
            if n:
                fb["fund_balances"].setdefault(kk, round(n, 2))
        for kk, vv in (o.get("bank") or {}).items():
            n = _num(vv)
            if n:
                fb["bank"][kk] = round(n, 2)
        rt = _num(o.get("receivables_total"))
        if rt:
            fb["receivables_total"] = max(fb["receivables_total"] or 0, round(rt, 2))
        for nn in o.get("notes", []) or []:
            if nn:
                fb["notes"].append(str(nn)[:160])

    # deterministic fallbacks
    fb["bank"].setdefault("general_fund_closing", fin.bank_general_closing)
    fb["bank"].setdefault("reserve_fund_closing", fin.bank_reserve_closing)
    if not fb["suppliers"] and fin.suppliers:
        fb["suppliers"] = [{"name": n, "balance": round(b, 2)} for n, b in fin.suppliers[:30]]
    return fb


# --------------------------------------------------------------------------- #
# 4. per-sheet mapping
# --------------------------------------------------------------------------- #
_MAP_SYSTEM = (
    "You are a senior UAE RERA audit working-paper preparer. You are given ONE "
    "annexure sheet's EXACT current layout (every non-empty cell, existing "
    "formulas verbatim) and a FACTBOOK of figures extracted from the client's "
    "trial balance, bank statements, budget and schedules.\n"
    "Populate EVERY empty data cell you can from the FACTBOOK:\n"
    "1. Never overwrite an existing =formula or a row/column label.\n"
    "2. Line-item amounts: pick the FACTBOOK figure whose label matches the "
    "template row; write it as a positive number.\n"
    "3. Every calculated cell MUST be an Excel formula - totals =SUM(...), "
    "sub-totals, differences =B5-C5, running balances, and cross-sheet references "
    "='Sheet Name'!K25 where the template implies it.\n"
    "4. Leave a cell blank ONLY if the FACTBOOK genuinely has nothing for it. "
    "Never invent a number.\n"
    "5. Stay within the sheet's used rows; keep values plain (number, or string "
    "starting with =, or short text).\n"
    'Return {"writes":[{"sheet":"<exact sheet name>","cell":"G13","value":<number|string>,"note":"<why, short>"}]}'
)


def _dump_sheet(ws) -> str:
    merged = [str(m) for m in ws.merged_cells.ranges][:20]
    out = [f'### SHEET: "{ws.title}"' + (f"   merged: {', '.join(merged)}" if merged else "")]
    n = 0
    for r in range(1, min(ws.max_row, 130) + 1):
        cells = []
        for c in range(1, min(ws.max_column, 18) + 1):
            v = ws.cell(row=r, column=c).value
            if v in (None, ""):
                continue
            s = str(v)
            cells.append(f"{get_column_letter(c)}{r}: {s[:58]}")
        if cells:
            out.append("  " + "  |  ".join(cells))
            n += 1
        if n >= _MAX_LAYOUT_ROWS:
            out.append("  ...(more rows)")
            break
    return "\n".join(out)


def _factbook_json(fb: dict) -> str:
    return json.dumps(fb, ensure_ascii=False, separators=(",", ":"))[:9000]


# --------------------------------------------------------------------------- #
# public
# --------------------------------------------------------------------------- #
def propose_writes(
    wb, ctx: EntityContext, fin: FinancialData, doc_paths: list[tuple[Path, str, str]] | None = None
) -> list[dict]:
    if not llm_client.available():
        return []
    doc_paths = doc_paths or []

    # 1-2: classify + extract
    fact_objs: list[dict] = []
    for path, fname, folder in doc_paths:
        kind = classify(fname, folder, Path(fname).suffix.lower())
        if kind in _FINANCIAL_KINDS and Path(path).exists():
            fact_objs.append(_extract_file(Path(path), kind, fname))
            time.sleep(_CALL_GAP_S)

    # 3: consolidate
    fb = consolidate(fact_objs, fin, ctx)
    fb_json = _factbook_json(fb)

    # 4: per-sheet mapping (2 sheets per call to stay under free-tier limits)
    out: list[dict] = []
    sheets = [ws for ws in wb.worksheets if ws.title.strip().lower() not in _SKIP_SHEETS]
    for i in range(0, len(sheets), 2):
        batch = sheets[i : i + 2]
        names = ", ".join(f'"{ws.title}"' for ws in batch)
        prompt = (
            f"FACTBOOK:\n{fb_json}\n\n"
            + "\n\n".join(_dump_sheet(ws) for ws in batch)
            + f"\n\nReturn the writes for sheets {names}. Every write MUST include the exact sheet name."
        )
        try:
            data = llm_client.chat_json(_MAP_SYSTEM, prompt, max_tokens=3200, retries=3)
            for wobj in data.get("writes", []):
                if isinstance(wobj, dict) and wobj.get("cell") and wobj.get("sheet"):
                    out.append(wobj)
        except Exception as exc:  # noqa: BLE001
            out.append({"__error__": f"{type(exc).__name__}: {exc}", "sheet": batch[0].title})
        time.sleep(_CALL_GAP_S)

    return out


# kept for backwards-compatible import in procedures.py
def data_pack(ctx: EntityContext, fin: FinancialData) -> str:
    lines = [f"ENTITY: {ctx.jop_name or 'n/a'} | period {ctx.period_start} to {ctx.period_end}"]
    for b in ("income_service_charge", "income_other", "exp_services", "exp_maintenance",
              "exp_utility", "exp_management_fee", "exp_insurance", "exp_reserve_fund",
              "liab_accounts_payable", "liab_accrued", "asset_sc_receivable"):
        v = fin.total(b)
        lines.append(f"  {b}: {v:,.2f}" if isinstance(v, (int, float)) else f"  {b}: not found")
    lines.append(f"  bank GF closing: {fin.bank_general_closing}")
    lines.append(f"  bank RF closing: {fin.bank_reserve_closing}")
    if fin.suppliers:
        lines.append("  suppliers: " + "; ".join(f"{n} {b:,.0f}" for n, b in fin.suppliers[:20]))
    return "\n".join(lines)
