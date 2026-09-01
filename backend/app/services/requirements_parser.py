"""Parse the auditor's *Requirements Checklist* workbook into structured line items.

Handles the two layouts seen in the wild:

**Layout A** - ref number in column A (often a formula ``=A19+0.1``), requirement
in B, evidence in C, STATUS in D, REMARKS in E; section banners are an integer in
A + an UPPER-CASE title in B.

**Layout B** ("New list ...") - the marker word ``Annexure`` sits in one column
with the section title beside it; each item is ``"1.1 RERA registration
certificate"`` (ref + text in a single cell), followed by a *Responsible
Department* column, then STATUS and REMARKS.

Workbooks frequently hold *several* requirement sheets (an old one and a new
one).  We score every sheet and parse the richest, and expose the full list so
the UI can let the auditor switch.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from ..schemas import RequirementItem, RequirementsResult

_STATUS_CANON = {
    "provided": "Received", "received": "Received", "yes": "Received", "done": "Received",
    "pending": "Pending", "not provided": "Pending", "outstanding": "Pending",
    "partial": "Partial", "partially provided": "Partial", "provided (partial)": "Partial",
    "n/a": "Not applicable", "na": "Not applicable", "not applicable": "Not applicable",
}
_STATUS_WORDS = set(_STATUS_CANON) | {"provided", "pending", "received", "partial", "n/a"}

_WS_RE = re.compile(r"\s+")
_DASHES = {"–": "-", "—": "-", "�": "-", "\xa0": " "}
# "1.1 RERA registration certificate"  ->  ("1.1", "RERA registration certificate")
# deliberately requires whitespace + a letter so a cached formula value like
# "1.2000000000000002" is NOT mistaken for a reference.
_INLINE_REF = re.compile(r"^\s*(\d{1,2}(?:\.\d{1,3}){1,3})[.)]?\s+([A-Za-z(\"'].*\S)$", re.S)
_SECTION_NUM = re.compile(r"^\s*(?:SECTION\s+)?(\d{1,2})\b", re.I)


def canon_status(raw) -> str:
    if raw is None:
        return "Pending"
    key = str(raw).strip().lower()
    if key in _STATUS_CANON:
        return _STATUS_CANON[key]
    return raw if isinstance(raw, str) and key else "Pending"


def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v)
    for bad, good in _DASHES.items():
        s = s.replace(bad, good)
    return _WS_RE.sub(" ", s).strip()


# --------------------------------------------------------------------------- #
# sheet selection
# --------------------------------------------------------------------------- #
def _score_sheet(ws) -> int:
    """Rough count of requirement-like rows on a sheet."""
    n = 0
    for r in range(1, min(ws.max_row, 400) + 1):
        row = [_clean(ws.cell(row=r, column=c).value) for c in range(1, min(ws.max_column, 8) + 1)]
        has_status = any(x.lower() in _STATUS_WORDS for x in row)
        has_ref = any(_INLINE_REF.match(x) for x in row)
        if has_status and (has_ref or any(len(x) > 12 for x in row)):
            n += 1
    return n


def _sheet_candidates(wb) -> list[str]:
    scored = sorted(
        ((s, _score_sheet(wb[s])) for s in wb.sheetnames),
        key=lambda kv: kv[1],
        reverse=True,
    )
    good = [s for s, sc in scored if sc >= 3]
    return good or [wb.sheetnames[0]]


# --------------------------------------------------------------------------- #
# table geometry
# --------------------------------------------------------------------------- #
def _detect_table(ws) -> dict:
    """Return {header_row, status, remarks, responsible?, evidence?, item, layout}."""
    for r in range(1, min(ws.max_row, 45) + 1):
        cells = {c: _clean(ws.cell(row=r, column=c).value) for c in range(1, min(ws.max_column, 12) + 1)}
        short = {c: t for c, t in cells.items() if 0 < len(t) <= 34}
        status_col = next(
            (c for c, t in short.items()
             if t.lower() == "status" or t.lower().endswith(" status") or t.lower().startswith("status")),
            None,
        )
        req_col = next(
            (c for c, t in short.items()
             if ("requirement" in t.lower() or "document" in t.lower()) and "type" not in t.lower() and "status" not in t.lower()),
            None,
        )
        if not (status_col and req_col and status_col != req_col):
            continue
        rem_col = next((c for c, t in short.items() if "remark" in t.lower() or "comment" in t.lower()), status_col + 1)
        ev_col = next((c for c, t in short.items() if "evidence" in t.lower() or ("document" in t.lower() and "type" in t.lower())), None)
        resp_col = next((c for c, t in short.items() if "responsible" in t.lower() or (t.lower().endswith("department") and c != req_col)), None)

        # which column actually holds the item text (ref + description)?
        item_col = _find_item_col(ws, r, status_col, [req_col, ev_col, req_col - 1, req_col + 1])
        return {
            "header_row": r,
            "status": status_col,
            "remarks": rem_col,
            "evidence": ev_col if ev_col and ev_col != item_col else None,
            "responsible": resp_col if resp_col and resp_col not in (item_col, status_col) else None,
            "item": item_col,
        }
    # fall back to the classic PROA layout
    return {"header_row": 18, "status": 4, "remarks": 5, "evidence": 3, "responsible": None, "item": 2}


def _find_item_col(ws, header_row: int, status_col: int, prefer: list[int | None]) -> int:
    """The column that holds the item description (with or without an inline ref)."""
    counts: dict[int, int] = {}
    for c in range(1, min(ws.max_column, 6) + 1):
        hit = 0
        for r in range(header_row + 1, min(ws.max_row, header_row + 120) + 1):
            t = _clean(ws.cell(row=r, column=c).value)
            if not t:
                continue
            st = _clean(ws.cell(row=r, column=status_col).value)
            has_alpha = any(ch.isalpha() for ch in t)
            is_number = re.fullmatch(r"[\d.,\-]+", t) is not None
            if _INLINE_REF.match(t):
                hit += 3
            elif is_number:
                hit -= 2
            elif has_alpha and (len(t) > 15 or (st and len(t) > 6)):
                hit += 1
        counts[c] = hit
    best = max(counts, key=counts.get)
    if counts[best] <= 0:
        for p in prefer:
            if p and 1 <= p <= ws.max_column:
                return p
        return 2
    return best


# --------------------------------------------------------------------------- #
def _section_of(cells: list[str], has_data: bool = True) -> tuple[int | None, str] | None:
    """(number, title) when the row is a section banner, else None."""
    low = [c.lower() for c in cells]
    if "annexure" in low:
        idx = low.index("annexure")
        title = next((c for c in cells[idx + 1:] if c), "") or next((c for c in cells if c and c.lower() != "annexure"), "")
        m = _SECTION_NUM.match(title)
        return (int(m.group(1)) if m else None), re.sub(r"^\s*\d+[.\s]*", "", title).strip() or title
    # Layout A: a bare section number + a title, and nothing else on the row
    for i, c in enumerate(cells[:2]):
        if re.fullmatch(r"\d{1,2}", c) and len(cells) > i + 1 and cells[i + 1]:
            title = cells[i + 1]
            letters = [ch for ch in title if ch.isalpha()]
            upper = letters and sum(ch.isupper() for ch in letters) / len(letters) > 0.6
            if upper or (not has_data and len(title) >= 10 and not _INLINE_REF.match(title)):
                m = _SECTION_NUM.match(title)
                num = int(m.group(1)) if m else int(c)
                return num, re.sub(r"^\s*SECTION\s+\d+\s*[:\-]?\s*", "", title, flags=re.I).strip()
    for c in cells[:2]:
        if re.match(r"^\s*SECTION\s+\d", c, re.I):
            m = re.search(r"(\d{1,2})", c)
            title = re.sub(r"^\s*SECTION\s+\d+\s*[:\-]?\s*", "", c, flags=re.I).strip()
            return (int(m.group(1)) if m else None), title or c
    return None


_PLACEHOLDER_CLIENT = re.compile(r"\b(oa\s*name|project\s*name|client\s*name|xxx+|<[^>]+>)\b", re.I)


def _is_placeholder_client(name: str) -> bool:
    n = _clean(name)
    return (not n) or bool(_PLACEHOLDER_CLIENT.search(n)) or len(n) < 4


def _preamble(ws, header_row: int) -> tuple[str, str]:
    client = period = ""
    for r in range(1, max(header_row, 2)):
        for c in range(1, min(ws.max_column, 7) + 1):
            t = _clean(ws.cell(row=r, column=c).value)
            low = t.lower()
            if not client and low.startswith("client"):
                client = re.sub(r"^client(\s*name)?\s*[:\-]?\s*", "", t, flags=re.I)
            if not period and "audit period" in low:
                period = re.sub(r"^audit period\s*[:\-]?\s*", "", t, flags=re.I)
    return client, period


def parse_requirements(path: str | Path, sheet: str | None = None) -> RequirementsResult:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)

    candidates = _sheet_candidates(wb)
    chosen = sheet if (sheet and sheet in wb.sheetnames) else candidates[0]
    ws = wb[chosen]

    geo = _detect_table(ws)
    hr = geo["header_row"]
    c_item, c_status, c_rem = geo["item"], geo["status"], geo["remarks"]
    c_ev, c_resp = geo["evidence"], geo["responsible"]

    client_name, audit_period = _preamble(ws, hr)
    # this sheet may carry only a placeholder ("OA Name_ Project Name") - borrow
    # the real client / period from a sibling requirements sheet
    if _is_placeholder_client(client_name) or not audit_period:
        for other in candidates:
            if other == chosen:
                continue
            og = _detect_table(wb[other])
            oc, op = _preamble(wb[other], og["header_row"])
            if _is_placeholder_client(client_name) and oc and not _is_placeholder_client(oc):
                client_name = oc
            if not audit_period and op:
                audit_period = op

    items: list[RequirementItem] = []
    section_titles: dict[int, str] = {}
    cur_section, cur_title, sub_index = 0, "General", 0

    for r in range(hr + 1, ws.max_row + 1):
        cells = [_clean(ws.cell(row=r, column=c).value) for c in range(1, min(ws.max_column, 10) + 1)]
        item_txt = cells[c_item - 1] if c_item - 1 < len(cells) else ""
        status_txt = cells[c_status - 1] if c_status - 1 < len(cells) else ""
        ev_txt = cells[c_ev - 1] if c_ev and c_ev - 1 < len(cells) else ""
        resp_txt = cells[c_resp - 1] if c_resp and c_resp - 1 < len(cells) else ""
        rem_txt = cells[c_rem - 1] if c_rem - 1 < len(cells) else ""

        has_data = bool(status_txt or ev_txt or resp_txt)
        sec = _section_of(cells, has_data)
        if sec is not None:
            num, title = sec
            cur_section = num if num else (cur_section + 1)
            cur_title = title or f"Section {cur_section}"
            section_titles[cur_section] = cur_title
            sub_index = 0
            continue

        if not item_txt:
            continue
        low_it = item_txt.lower()
        if low_it.startswith((
            "should you have any queries", "yours sincerely", "on behalf of",
            "parker russell", "member of parker", "audit manager", "audit partner",
        )):
            break  # sign-off block - table is over
        if low_it.startswith(("kindly update", "to commence", "requirements for the audit")):
            continue
        # post-table prose / addresses: no status, no evidence, no inline ref
        if not has_data and not _INLINE_REF.match(item_txt):
            continue

        m = _INLINE_REF.match(item_txt)
        if m and m.group(2):
            ref, requirement = m.group(1), m.group(2).strip(" .:-)")
        else:
            sub_index += 1
            ref = f"{cur_section}.{sub_index}" if cur_section else f"0.{sub_index}"
            requirement = item_txt
        if m and m.group(2):
            # keep sub_index roughly in step for later fallbacks
            tail = m.group(1).split(".")
            if len(tail) >= 2 and tail[-1].isdigit():
                sub_index = max(sub_index, int(tail[-1]))

        items.append(RequirementItem(
            ref=ref,
            section=cur_section,
            section_title=cur_title,
            requirement=requirement or item_txt,
            evidence_type=ev_txt,
            responsible=resp_txt,
            sheet_status=canon_status(status_txt),
            remarks=rem_txt,
            row=r,
        ))

    wb.close()
    return RequirementsResult(
        filename=path.name,
        sheet=chosen,
        available_sheets=candidates,
        client_name=client_name,
        audit_period=audit_period,
        items=items,
        section_titles=section_titles,
    )


# kept for import compatibility (requirements_filler uses it)
def _find_table(ws):
    g = _detect_table(ws)
    return (
        g["header_row"], max(1, g["item"] - 1), g["item"],
        g["evidence"] or g["item"] + 1, g["status"], g["remarks"],
    )
