"""Populate a copy of the master template while preserving all formatting.

Strategy: open the template with ``openpyxl`` (keeps styles, fonts, fills, number
formats, merged cells, column widths).  We only ever *set ``cell.value``* on
cells that already exist - never restyle - except for **status cells**, whose
left-over sample fills are deliberately overwritten with a fill that matches the
status word so the colour coding is correct.

Calculated cells are written as **Excel formula strings** so they recalculate
natively in Microsoft Excel.  Every write is recorded in a :class:`WorkbookReport`.
"""
from __future__ import annotations

import re
from copy import copy
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..config import LLM_WORKBOOK
from ..schemas import CellWrite, EntityContext, MatchResult, RequirementsResult, WorkbookReport
from . import llm_client, workbook_llm
from .financials import FinancialData

_CELL_RE = re.compile(r"^([A-Za-z]{1,3})(\d{1,5})$")
_NUM_RE = re.compile(r"^-?\(?\d[\d,]*\.?\d*\)?%?$")
# a genuine fill-in-the-blank placeholder (NOT a "Name:" style label)
_PLACEHOLDER_RE = re.compile(r"_{3,}|x{4,}|\.{4,}|<[^>]*>|\[[^\]]*\]", re.I)


def _xref(title: str, cell: str) -> str:
    """Excel cross-sheet reference with the sheet name safely quoted."""
    return f"='{title.replace(chr(39), chr(39) * 2)}'!{cell}"

# --------------------------------------------------------------------------- #
# status colour coding (standard Excel Good / Neutral / Bad palette)
# --------------------------------------------------------------------------- #
_STATUS_FILL = {
    "Received": PatternFill("solid", fgColor="C6EFCE"),
    "Partial": PatternFill("solid", fgColor="FFEB9C"),
    "Pending": PatternFill("solid", fgColor="FFC7CE"),
    "Not applicable": PatternFill("solid", fgColor="E7E6E6"),
}
_STATUS_FONT_RGB = {
    "Received": "FF006100",
    "Partial": "FF9C5700",
    "Pending": "FF9C0006",
    "Not applicable": "FF3F3F3F",
}
_NO_FILL = PatternFill(fill_type=None)

# entity-name placeholders used throughout the sample template (longest first)
_ENTITY_PLACEHOLDERS = [
    "jumeirah living - world trade centre residences",
    "jumeirah living – world trade centre residences",
    "jumeirah living — world trade centre residences",
    "world trade centre residences",
    "world trade center residences",
    "jumeirah living",
    "park ville 07",
    "park ville",
]


def _norm(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", name.lower())


def _is_merged_child(cell) -> bool:
    return type(cell).__name__ == "MergedCell"


class _Writer:
    def __init__(self, wb):
        self.wb = wb
        self.writes: list[CellWrite] = []
        self.formulas = 0
        self.warnings: list[str] = []
        self.sheets_touched: set[str] = set()
        self.links: dict[str, str] = {}     # logical name -> "'Sheet'!A1" ref cell

    def sheet(self, *candidates: str):
        want = [_norm(c) for c in candidates]
        for ws in self.wb.worksheets:
            n = _norm(ws.title)
            if any(n == w or n.startswith(w) or w in n for w in want):
                return ws
        return None

    @staticmethod
    def _anchor(ws, row: int, col: int) -> tuple[int, int]:
        for rng in ws.merged_cells.ranges:
            if (rng.min_row <= row <= rng.max_row) and (rng.min_col <= col <= rng.max_col):
                return rng.min_row, rng.min_col
        return row, col

    def put(self, ws, row: int, col: int, value, source: str, inherit: bool = False):
        row, col = self._anchor(ws, row, col)
        cell = ws.cell(row=row, column=col)
        if isinstance(value, str) and value.startswith("="):
            self.formulas += 1
        cell.value = value
        if inherit:
            self.inherit_style(ws, row, col)
        self.writes.append(
            CellWrite(sheet=ws.title, cell=f"{get_column_letter(col)}{row}", value=str(value), source=source)
        )
        self.sheets_touched.add(ws.title)
        return cell

    @staticmethod
    def inherit_style(ws, row: int, col: int) -> None:
        """Give a freshly-written cell the look of the styled data rows above it
        (borders / fill / font / number format) so appended rows match the table."""
        cell = ws.cell(row=row, column=col)
        b = cell.border
        if b and (b.left.style or b.right.style or b.bottom.style or b.top.style):
            return  # already inside the formatted band
        for up in range(row - 1, max(0, row - 8), -1):
            src = ws.cell(row=up, column=col)
            sb = src.border
            if (sb and (sb.left.style or sb.bottom.style or sb.top.style)) or (
                src.fill and src.fill.patternType
            ):
                cell.font = copy(src.font)
                cell.border = copy(src.border)
                cell.fill = copy(src.fill)
                cell.alignment = copy(src.alignment)
                cell.number_format = src.number_format
                return

    def put_status(self, ws, row: int, col: int, status: str, source: str):
        cell = self.put(ws, row, col, status, source)
        fill = _STATUS_FILL.get(status)
        if fill:
            cell.fill = fill
            old = cell.font
            cell.font = Font(
                name=old.name, size=old.size, bold=old.bold, italic=old.italic,
                color=_STATUS_FONT_RGB[status],
            )

    def clear_fill(self, ws, row: int, col: int):
        r, c = self._anchor(ws, row, col)
        ws.cell(row=r, column=c).fill = _NO_FILL

    def put_formula_if_blank(self, ws, row: int, col: int, formula: str, source: str):
        ar, ac = self._anchor(ws, row, col)
        cur = ws.cell(row=ar, column=ac).value
        if isinstance(cur, str) and cur.startswith("="):
            return
        self.put(ws, row, col, formula, source)

    @staticmethod
    def col_of_label(ws, *needles: str, max_row=8, max_col=25, min_col=1, maxlen=None) -> int | None:
        """Column index of the first header cell containing one of ``needles``.
        ``min_col`` skips left-hand label columns; ``maxlen`` rejects long title
        cells like "Annexure 10 : Bank Balances"."""
        needles = tuple(n.lower() for n in needles)
        for r in range(1, min(ws.max_row, max_row) + 1):
            for c in range(min_col, min(ws.max_column, max_col) + 1):
                v = ws.cell(row=r, column=c).value
                if not isinstance(v, str):
                    continue
                if maxlen is not None and len(v.strip()) > maxlen:
                    continue
                if any(n in v.lower() for n in needles):
                    return c
        return None

    @staticmethod
    def row_of_label(ws, *needles: str, col=2, max_row=140, start=1, maxlen=None) -> int | None:
        needles = tuple(n.lower() for n in needles)
        for r in range(start, min(ws.max_row, max_row) + 1):
            v = ws.cell(row=r, column=col).value
            if not isinstance(v, str):
                continue
            if maxlen is not None and len(v.strip()) > maxlen:
                continue
            if any(n in v.lower() for n in needles):
                return r
        return None


# --------------------------------------------------------------------------- #
# period / entity context
# --------------------------------------------------------------------------- #
class _Period:
    def __init__(self, project_name: str, period_end: str):
        self.project = project_name.strip() or "The JOP"
        self.end_dt = _parse_date(period_end) or date(date.today().year, 6, 30)
        self.year = self.end_dt.year
        self.prior_year = self.year - 1
        self.end_long = self.end_dt.strftime("%d %B %Y").lstrip("0")
        self.end_short = self.end_dt.strftime("%d %b %Y").lstrip("0")
        self.end_us = self.end_dt.strftime("%B %d, %Y")
        self.start_long = date(self.year, 1, 1).strftime("1 January %Y")
        self.start_us = date(self.year, 1, 1).strftime("January 1, %Y")


def _parse_date(s: str | None) -> date | None:
    if not s:
        return None
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %B %Y", "%d %b %Y", "%B %d, %Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.search(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", s)
    if m:
        try:
            return datetime.strptime(f"{m.group(1)} {m.group(2)} {m.group(3)}", "%d %B %Y").date()
        except ValueError:
            pass
    m = re.search(r"(20\d{2})", s)
    if m:
        return date(int(m.group(1)), 6, 30)
    return None


def derive_period(requirements: RequirementsResult | None, project_name: str | None, period_end: str | None) -> _Period:
    name = project_name or (requirements.client_name if requirements else "") or "The JOP"
    pe = period_end
    if not pe and requirements:
        pe = requirements.audit_period or requirements.filename
    return _Period(name, pe or "")


# --------------------------------------------------------------------------- #
# global pass: entity name + reporting period across every sheet
# --------------------------------------------------------------------------- #
# cells where a bare 4-digit year should be rolled forward to the reporting
# period.  Deliberately excludes "year" so historical buckets such as
# "Accrued in year 2022" are left untouched.
_LABELISH = re.compile(
    r"jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|as at|as of|audited|"
    r"balance|period|annexure|ended|ending|statement",
    re.I,
)


def _swap_dates(text: str, p: _Period) -> str:
    t = text
    t = re.sub(r"\b30\s+June\s+2023\b", f"30 June {p.year}", t, flags=re.I)
    t = re.sub(r"\b30\s+Jun\s+2023\b", f"30 Jun {p.year}", t, flags=re.I)
    t = re.sub(r"\bJune\s+30,\s*2023\b", f"June 30, {p.year}", t, flags=re.I)
    t = re.sub(r"\b1\s+Jan(uary)?\s+2023\b", p.start_long, t, flags=re.I)
    t = re.sub(r"\bJanuary\s+1,\s*2023\b", p.start_us, t, flags=re.I)
    t = re.sub(r"\b1 Jan to 30 Jun 2023\b", f"1 Jan to 30 Jun {p.year}", t, flags=re.I)
    if _LABELISH.search(t):
        t = re.sub(r"\b2023\b", str(p.year), t)
        t = re.sub(r"\b2022\b", str(p.prior_year), t)
    return t


def _apply_entity_and_period(w: _Writer, p: _Period) -> None:
    changed = 0
    for ws in w.wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.strip() or _is_merged_child(cell):
                    continue
                if v.startswith("="):        # never rewrite a formula
                    continue
                was_upper = v.isupper()
                new = v
                for ph in _ENTITY_PLACEHOLDERS:
                    if ph in new.lower():
                        new = re.sub(re.escape(ph), p.project, new, flags=re.I)
                # collapse "NAME - NAME" left behind when a compound placeholder
                # was replaced piece-by-piece
                dup = re.escape(p.project)
                new = re.sub(rf"{dup}\s*[-–—]\s*{dup}", p.project, new, flags=re.I)
                new = _swap_dates(new, p)
                if was_upper and new != v:
                    new = new.upper()
                if new != v:
                    cell.value = new
                    changed += 1
    if changed:
        w.sheets_touched.add("(entity & period)")
        w.warnings.append(f"entity/period: updated {changed} label cell(s) to '{p.project}' / {p.year}.")


# --------------------------------------------------------------------------- #
# Sheet 17 - PBC Checklist Status
# --------------------------------------------------------------------------- #
_REF_RE = re.compile(r"^\d{1,2}\.\d{1,3}(\.\d{1,3})?$")


def _fill_pbc_checklist(w: _Writer, match: MatchResult, reqs: RequirementsResult | None) -> None:
    ws = w.sheet("17. PBC Checklist Status", "17 PBC", "PBC Checklist")
    if ws is None:
        w.warnings.append("Sheet '17. PBC Checklist Status' not found - skipped.")
        return
    by_ref = {r.ref: r for r in match.rows}
    req_by_ref = {i.ref: i for i in (reqs.items if reqs else [])}
    first_row = last_row = None
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a is None or not _REF_RE.match(str(a).strip()):
            continue
        key = str(a).strip()
        first_row = first_row or r
        last_row = r
        w.clear_fill(ws, r, 2)  # drop stray yellow highlight
        req = req_by_ref.get(key)
        if req and not ws.cell(row=r, column=2).value:
            w.put(ws, r, 2, req.requirement[:200], "requirements")
        row = by_ref.get(key)
        if row is None:
            w.put_status(ws, r, 3, "Pending", "requirements")
            w.put(ws, r, 4, "Not in matched set.", "requirements")
            continue
        w.put_status(ws, r, 3, row.status, "match")
        remark = row.comment or row.llm_rationale or (
            "Document(s): " + ", ".join(m.filename for m in row.matched_files)
            if row.matched_files else "Not yet provided by client."
        )
        w.put(ws, r, 4, remark[:250], "match")

    if first_row and last_row:
        rng = f"C{first_row}:C{last_row}"
        for label, status in [
            ("Provided", "Received"), ("Partially Provided", "Partial"),
            ("Not Provided", "Pending"), ("Not Applicable", "Not applicable"),
        ]:
            sr = w.row_of_label(ws, label, col=1, max_row=ws.max_row)
            if sr:
                w.put(ws, sr, 2, f'=COUNTIF({rng},"{status}")', "formula")
        tr = w.row_of_label(ws, "total requirements", col=1, max_row=ws.max_row)
        if tr:
            w.put(ws, tr, 2, f"=COUNTA({rng})", "formula")


# --------------------------------------------------------------------------- #
# Sheet 18 - Requirements Tracker
# --------------------------------------------------------------------------- #
def _fill_requirements_tracker(w: _Writer, match: MatchResult, reqs: RequirementsResult | None) -> None:
    ws = w.sheet("18. Requirements Tracker", "18 Requirements", "Requirements Tracker")
    if ws is None:
        w.warnings.append("Sheet '18. Requirements Tracker' not found - skipped.")
        return
    by_ref = {r.ref: r for r in match.rows}
    req_by_ref = {i.ref: i for i in (reqs.items if reqs else [])}
    section_titles = reqs.section_titles if reqs else {}
    sec_re = re.compile(r"^section\s+(\d{1,2})$", re.I)
    today = date.today().strftime("%Y-%m-%d")
    first_row = last_row = None
    for r in range(2, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a is None:
            continue
        key = str(a).strip()
        sm = sec_re.match(key)
        if sm:
            title = section_titles.get(int(sm.group(1)))
            if title and not ws.cell(row=r, column=2).value:
                w.put(ws, r, 2, title, "requirements")
            continue
        if not _REF_RE.match(key):
            continue
        first_row = first_row or r
        last_row = r
        row = by_ref.get(key)
        req = req_by_ref.get(key)
        if req and not ws.cell(row=r, column=2).value:
            w.put(ws, r, 2, req.requirement[:200], "requirements")
        if row is None:
            continue
        w.put_status(ws, r, 3, row.status, "match")
        w.put(ws, r, 4, "Yes" if row.status in ("Pending", "Partial") else "No", "match")
        w.put(ws, r, 5, today if row.status == "Received" else "", "match")
        note = row.comment or row.llm_rationale or ", ".join(m.filename for m in row.matched_files)
        if note:
            w.put(ws, r, 6, note[:250], "match")

    pr = w.row_of_label(ws, "progress", col=1, max_row=12, maxlen=20)
    if pr and first_row:
        rng = f"C{first_row}:C{last_row}"
        w.put(
            ws, pr, 2,
            f'=COUNTIF({rng},"Received")&" of "&COUNTA({rng})&" requirements received; "'
            f'&COUNTIF({rng},"Pending")&" pending, "&COUNTIF({rng},"Partial")&" partial"',
            "formula",
        )


# --------------------------------------------------------------------------- #
# value column detection for the financial annexures
# --------------------------------------------------------------------------- #
def _value_col(ws, default: int) -> int:
    c = _Writer.col_of_label(ws, "1 jan", "period ended", "as at", "30 jun", "6 months", "balance 30", max_row=14, min_col=3, maxlen=32)
    return c or default


def _sum_between(w: _Writer, ws, col: int, top_label_row: int, total_row: int, source: str):
    if top_label_row and total_row and total_row > top_label_row + 1:
        letter = get_column_letter(col)
        w.put_formula_if_blank(
            ws, total_row, col, f"=SUM({letter}{top_label_row + 1}:{letter}{total_row - 1})", source
        )


# --------------------------------------------------------------------------- #
# Sheet 2 - Income Statement
# --------------------------------------------------------------------------- #
_IS_MAP = [
    (("service charge",), "income_service_charge"),
    (("chilled water", "in-unit"), "income_chilled_water"),
    (("other income",), "income_other"),
    (("services",), "exp_services"),
    (("maintenance",), "exp_maintenance"),
    (("community improvement",), "exp_community_improvement"),
    (("utility",), "exp_utility"),
    (("management fee",), "exp_management_fee"),
    (("insurance",), "exp_insurance"),
    (("master community",), "exp_master_community"),
    (("provision for expected credit", "expected credit loss"), "exp_provision_ecl"),
    (("reserve fund expense",), "exp_reserve_fund"),
]


def _fill_income_statement(w: _Writer, fin: FinancialData) -> None:
    ws = w.sheet("2 _IS", "2_IS", "2 IS", "income statement")
    if ws is None:
        w.warnings.append("Sheet '2 _IS' not found - skipped.")
        return
    col = _value_col(ws, default=7)
    # line item -> cross-sheet link (preferred over a hard number)
    line_links = {
        "income_other": "other_income_recognized",
    }
    wrote = 0
    for needles, bucket in _IS_MAP:
        row = w.row_of_label(ws, *needles, col=2, max_row=ws.max_row)
        if not row:
            continue
        link = w.links.get(line_links.get(bucket, ""))
        if link:
            w.put(ws, row, col, link, f"link -> {bucket}")
            wrote += 1
            continue
        val = fin.total(bucket)
        if val is not None:
            w.put(ws, row, col, float(val), f"trial balance ({bucket})")
            wrote += 1
    ti = w.row_of_label(ws, "total income", col=2)
    inc_hdr = w.row_of_label(ws, "income", col=2)
    if inc_hdr and ti:
        _sum_between(w, ws, col, inc_hdr, ti, "formula")
    te = w.row_of_label(ws, "total expenditure", col=2)
    exp_hdr = w.row_of_label(ws, "expenditure", col=2)
    if exp_hdr and te:
        _sum_between(w, ws, col, exp_hdr, te, "formula")
    sd = w.row_of_label(ws, "surplus", "deficit", col=2)
    if sd and ti and te:
        letter = get_column_letter(col)
        w.put_formula_if_blank(ws, sd, col, f"={letter}{ti}-{letter}{te}", "formula")
    if not wrote:
        w.warnings.append("Sheet '2 _IS': no trial-balance figures matched income/expense lines.")


# --------------------------------------------------------------------------- #
# Sheet 3 - Balance Sheet
# --------------------------------------------------------------------------- #
_BS_ASSET_MAP = [
    (("security deposit",), "asset_deposits"),
    (("prepaid",), "asset_prepaid"),
    (("service charge receivable",), "asset_sc_receivable"),
    (("provision for expected credit",), "asset_provision_ecl"),
    (("short term deposit", "short-term deposit"), "asset_short_term_deposit"),
    (("regulated bank - reserve", "reserve fund"), "asset_bank_reserve"),
    (("regulated bank - general", "general fund"), "asset_bank_general"),
]
_BS_LIAB_MAP = [
    (("security deposit",), "liab_security_deposit"),
    (("received in advance", "advance"), "liab_sc_advance"),
    (("accounts payable",), "liab_accounts_payable"),
    (("accrued expense", "accrued"), "liab_accrued"),
]


def _fill_balance_sheet(w: _Writer, fin: FinancialData, p: "_Period") -> None:
    ws = w.sheet("3 BS", "3_BS", "balance sheet", "financial position")
    if ws is None:
        w.warnings.append("Sheet '3 BS' not found - skipped.")
        return
    col = _value_col(ws, default=11)

    # refresh the "As at <date>" column headers (stored as real dates)
    for r in range(8, 14):
        for c in range(col - 1, min(ws.max_column, col + 3) + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, datetime):
                new = p.end_dt if c <= col else date(p.prior_year, 12, 31)
                w.put(ws, r, c, datetime(new.year, new.month, new.day), "period")
    # prefer a live reference to the supporting annexure over a static number
    bucket_link = {
        "asset_bank_general": "bank_gf",
        "asset_bank_reserve": "bank_rf",
        "liab_accounts_payable": "ap_total",
        "liab_accrued": "accrual_total",
    }

    def place(row: int, bucket: str, value) -> None:
        link = w.links.get(bucket_link.get(bucket, ""))
        if link:
            w.put(ws, row, col, link, f"link -> {bucket}")
        elif value is not None:
            w.put(ws, row, col, float(value), f"financials ({bucket})")

    def bank(bucket):
        if bucket == "asset_bank_general" and fin.bank_general_closing is not None:
            return fin.bank_general_closing
        if bucket == "asset_bank_reserve" and fin.bank_reserve_closing is not None:
            return fin.bank_reserve_closing
        return fin.total(bucket)

    seen: set[int] = set()
    for needles, bucket in _BS_ASSET_MAP:
        row = w.row_of_label(ws, *needles, col=2, max_row=ws.max_row)
        if not row or row in seen:
            continue
        val = bank(bucket)
        if w.links.get(bucket_link.get(bucket, "")) or val is not None:
            place(row, bucket, val)
            seen.add(row)

    liab_hdr = w.row_of_label(ws, "funds and liabilities", "liabilities", col=2)
    for needles, bucket in _BS_LIAB_MAP:
        row = w.row_of_label(ws, *needles, col=2, max_row=ws.max_row)
        if not row or row in seen or (liab_hdr and row < liab_hdr):
            continue
        val = fin.total(bucket)
        if w.links.get(bucket_link.get(bucket, "")) or val is not None:
            place(row, bucket, val)
            seen.add(row)

    assets_hdr = w.row_of_label(ws, "assets", col=2)
    total_assets = w.row_of_label(ws, "total assets", col=2)
    if assets_hdr and total_assets:
        _sum_between(w, ws, col, assets_hdr, total_assets, "formula")
    funds_hdr = w.row_of_label(ws, "accumulated funds", col=2)
    total_fl = w.row_of_label(ws, "total funds and liabilities", col=2)
    if funds_hdr and total_fl:
        _sum_between(w, ws, col, funds_hdr, total_fl, "formula")


# --------------------------------------------------------------------------- #
# Sheet 1 - Project details
# --------------------------------------------------------------------------- #
def _fill_project_details(w: _Writer, ctx: EntityContext, p: _Period) -> None:
    ws = w.sheet("1_project details", "1 project details", "project details")
    if ws is None:
        return
    val_col = 4
    q = {}
    for r in range(1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if isinstance(b, str) and b.strip():
            low = b.lower()
            if "name of the jop" in low:
                q["jop"] = r
            elif "developer name" in low:
                q["developer"] = r
            elif "registered address" in low:
                q["address"] = r
            elif "management company" in low:
                q["mc"] = r

    if "jop" in q:
        w.put(ws, q["jop"], val_col, ctx.jop_name or p.project, "entity")

    if "developer" in q:
        dev = ctx.developer_name
        if dev:
            bits = [dev]
            if ctx.developer_license:
                bits.append(f"Trade Licence No. {ctx.developer_license}")
            if ctx.developer_license_expiry:
                bits.append(f"expiry {ctx.developer_license_expiry}")
            w.put(ws, q["developer"], val_col, "; ".join(bits), "customer files")
        else:
            w.warnings.append("Sheet '1_project details': developer details not found in the customer files.")

    if "address" in q:
        if ctx.developer_address:
            w.put(ws, q["address"], val_col, ctx.developer_address, "customer files")
        else:
            w.warnings.append("Sheet '1_project details': developer address not found in the customer files.")

    if "mc" in q and ctx.management_company:
        mc = ctx.management_company
        if ctx.management_company_license:
            mc = f"{mc}; Trade Licence No. {ctx.management_company_license}"
        w.put(ws, q["mc"], val_col, mc, "customer files / requirements")


def _fill_signatory(w: _Writer, ctx: EntityContext, explicit: bool) -> None:
    """Fill the 'Name:' line under every 'Authorized Signatory' block.

    Only when a preparer name was explicitly supplied - otherwise the template's
    blank signature line is left for a hand signature.
    """
    if not explicit:
        return
    who = ctx.prepared_by
    if not who:
        return
    n = 0
    for ws in w.wb.worksheets:
        for r in range(1, ws.max_row + 1):
            for c in range(1, min(ws.max_column, 6) + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip().lower() in ("name:", "name :"):
                    above = ws.cell(row=r - 1, column=c).value
                    if isinstance(above, str) and "signatory" in above.lower():
                        w.put(ws, r, c, f"Name: {who}", "entity")
                        n += 1
    if n:
        w.warnings.append(f"signatory: filled {n} 'Name:' line(s) with '{who}'.")


# --------------------------------------------------------------------------- #
# Sheet 10 / 12 / 13 - bank, supplier & accrual balances
# --------------------------------------------------------------------------- #
def _fill_bank_balances(w: _Writer, fin: FinancialData) -> None:
    ws = w.sheet("10. bank balances", "10 bank balances", "bank balances")
    if ws is None:
        return
    bal_col = _Writer.col_of_label(ws, "balance", max_row=14, min_col=3, maxlen=18) or 11
    letter = get_column_letter(bal_col)
    gf_row = w.row_of_label(ws, "bank balance as of", "bank book balance", col=2, max_row=45)
    if gf_row and fin.bank_general_closing is not None:
        w.put(ws, gf_row, bal_col, float(fin.bank_general_closing), "bank statement (GF)")
        w.links["bank_gf"] = _xref(ws.title, f"{letter}{gf_row}")
    rf_row = w.row_of_label(ws, "bank balance as of", "accumulated reserve fund balance", col=2, start=46)
    if rf_row and fin.bank_reserve_closing is not None:
        w.put(ws, rf_row, bal_col, float(fin.bank_reserve_closing), "bank statement (RF)")
        w.links["bank_rf"] = _xref(ws.title, f"{letter}{rf_row}")
    if fin.bank_general_closing is None and fin.bank_reserve_closing is None:
        w.warnings.append("Sheet '10. bank balances': no bank closing balances detected.")


def _fill_supplier_balances(w: _Writer, fin: FinancialData) -> None:
    ws = w.sheet("12. supplier balance", "12 supplier balance", "supplier balance", "suppliers")
    if ws is None or not fin.suppliers:
        if ws is not None:
            w.warnings.append("Sheet '12. supplier balance': no supplier balances extracted.")
        return
    hdr = w.row_of_label(ws, "supplier name", col=2, max_row=ws.max_row)
    if not hdr:
        return
    bal_col = _Writer.col_of_label(ws, "closing balance", "ledger closing", max_row=hdr + 1, min_col=3, maxlen=32) or 7
    r = hdr + 1
    written = 0
    for name, bal in fin.suppliers[:30]:
        while r <= ws.max_row and ws.cell(row=r, column=2).value not in (None, ""):
            r += 1
        if r > ws.max_row:
            break
        w.put(ws, r, 2, name, "supplier ledger", inherit=True)
        w.put(ws, r, bal_col, float(bal), "supplier ledger", inherit=True)
        r += 1
        written += 1
    if written:
        letter = get_column_letter(bal_col)
        total_row = r + 1
        w.put(ws, total_row, bal_col, f"=SUM({letter}{hdr + 1}:{letter}{r})", "formula")
        w.links["ap_total"] = _xref(ws.title, f"{letter}{total_row}")


def _accrued_amount_col(ws, hdr: int) -> int | None:
    """The current-period 'Accrued in ...' column.

    The template marks it with a real *date* in the sub-header (hdr+1) while the
    historical buckets carry text like "Year 2021".  Prefer the dated column;
    fall back to the largest year number.
    """
    dated: list[tuple[datetime, int]] = []
    yeared: list[tuple[int, int]] = []
    for c in range(3, min(ws.max_column, 20) + 1):
        h = ws.cell(row=hdr, column=c).value
        if not (isinstance(h, str) and "accrued" in h.lower()):
            continue
        sub = ws.cell(row=hdr + 1, column=c).value
        if isinstance(sub, datetime):
            dated.append((sub, c))
        else:
            m = re.search(r"(19|20)\d{2}", str(sub or ""))
            yeared.append((int(m.group(0)) if m else 0, c))
    if dated:
        return max(dated)[1]
    if yeared:
        return max(yeared)[1]
    return None


def _fill_accruals(w: _Writer, fin: FinancialData) -> None:
    ws = w.sheet("13 accruals", "13accruals", "accrued expenses")
    if ws is None:
        return
    hdr = w.row_of_label(ws, "supplier name", col=2, max_row=ws.max_row)
    total_row = w.row_of_label(ws, "total", col=2, max_row=ws.max_row)
    if not hdr or not total_row:
        return
    amount_col = _accrued_amount_col(ws, hdr)
    if not amount_col:
        return
    letter = get_column_letter(amount_col)
    acc_total = fin.total("liab_accrued")

    if acc_total is not None:
        # append right after the last populated data row (keeps within the
        # template's existing SUM range)
        last = total_row - 1
        while last > hdr + 1 and not (
            ws.cell(row=last, column=2).value or ws.cell(row=last, column=5).value
        ):
            last -= 1
        target = last + 1 if last + 1 < total_row else last
        w.put(ws, target, 2, "Per trial balance (accrued expenses)", "trial balance", inherit=True)
        w.put(ws, target, amount_col, float(acc_total), "trial balance (liab_accrued)", inherit=True)

    w.put_formula_if_blank(
        ws, total_row, amount_col, f"=SUM({letter}{hdr + 2}:{letter}{total_row - 1})", "formula"
    )
    if acc_total is not None:
        w.links["accrual_total"] = _xref(ws.title, f"{letter}{total_row}")


# --------------------------------------------------------------------------- #
# Sheet 4 - Other income
# --------------------------------------------------------------------------- #
def _fill_other_income(w: _Writer, fin: FinancialData) -> None:
    ws = w.sheet("4_other income", "4 other income", "other income")
    if ws is None:
        return
    tr = w.row_of_label(ws, "total", col=2, max_row=ws.max_row)
    first_data = w.row_of_label(ws, "opening", col=2, max_row=ws.max_row)
    first_data = (first_data + 1) if first_data else 14
    rec_col = _Writer.col_of_label(ws, "recognized in pl", "recognized in p", max_row=12, min_col=3, maxlen=40)
    billed_col = _Writer.col_of_label(ws, "billed", max_row=12, min_col=3, maxlen=40)

    other = fin.total("income_other")
    chilled = fin.total("income_chilled_water")
    wrote = False
    if rec_col and other is not None:
        w.put(ws, first_data, 2, "Other income (per trial balance)", "trial balance", inherit=True)
        w.put(ws, first_data, rec_col, float(other), "trial balance (income_other)", inherit=True)
        wrote = True
        if billed_col:
            w.put(ws, first_data, billed_col, float(other), "trial balance (income_other)")
    if rec_col and chilled is not None:
        w.put(ws, first_data + 1, 2, "In-unit chilled water", "trial balance", inherit=True)
        w.put(ws, first_data + 1, rec_col, float(chilled), "trial balance (income_chilled_water)", inherit=True)
        wrote = True

    if tr and rec_col and wrote:
        w.links["other_income_recognized"] = _xref(ws.title, f"{get_column_letter(rec_col)}{tr}")
    if not wrote:
        w.warnings.append("Sheet '4_other income': no other-income figures available.")


# --------------------------------------------------------------------------- #
# Sheet 16 - Notes Summary
# --------------------------------------------------------------------------- #
def _fill_notes_summary(w: _Writer, match: MatchResult, fin: FinancialData) -> None:
    ws = w.sheet("16. Notes Summary", "16 notes summary", "notes summary")
    if ws is None:
        return
    hdr = w.row_of_label(ws, "summary of note", "no.", col=1, max_row=12) or w.row_of_label(
        ws, "no.", col=1, max_row=12
    )
    if not hdr:
        return
    notes: list[tuple[str, str, str]] = []
    for r in match.rows:
        if r.status == "Pending":
            notes.append(("17. PBC / 18. Tracker", "Scope - not provided",
                          f"[{r.ref}] {r.requirement[:110]} - not provided by the client."))
        elif r.status == "Partial":
            notes.append(("17. PBC / 18. Tracker", "Clarification requested",
                          f"[{r.ref}] {r.requirement[:90]} - document on file appears incomplete; confirmation requested."))
    for n in fin.notes:
        notes.append(("Financial annexures", "Data source", n[:150]))

    r = hdr + 1
    n = 0
    for annex, cat, summary in notes:
        if r > ws.max_row:
            break
        n += 1
        w.put(ws, r, 1, n, "notes")
        w.put(ws, r, 2, annex, "notes")
        w.put(ws, r, 4, cat, "notes")
        w.put(ws, r, 5, summary, "notes")
        r += 1
    if n:
        w.warnings.append(f"Sheet '16. Notes Summary': wrote {n} note(s).")


# --------------------------------------------------------------------------- #
# Audit Report - Procedures & Findings table
# --------------------------------------------------------------------------- #
def _boxrow(ws, r: int, border, c1: int = 2, c2: int = 10) -> None:
    for c in range(c1, c2 + 1):
        ws.cell(row=r, column=c).border = border


def _fill_audit_report_procedures(
    w: _Writer, fin: FinancialData, ctx: EntityContext, match: MatchResult, use_llm: bool
) -> None:
    ws = w.sheet("Audit Report", "audit report")
    if ws is None:
        return
    from .procedures import build_matrix

    anchor = w.row_of_label(ws, "the procedures we performed", "key findings", col=2, max_row=60)
    if not anchor:
        w.warnings.append("Audit Report: procedures anchor not found - table skipped.")
        return
    start = anchor + 2
    stop = w.row_of_label(ws, "summary of finding", col=2, start=start, max_row=ws.max_row) or (start + 140)
    stop = min(stop, ws.max_row)

    try:
        blocks, llm_used = build_matrix(fin, ctx, match, use_llm=use_llm)
    except Exception as exc:  # noqa: BLE001
        w.warnings.append(f"Audit Report: procedures matrix failed ({exc}).")
        return

    kf = w.row_of_label(ws, "key findings", col=2, max_row=start)
    if kf:
        w.put(ws, kf, 2, "Procedures and Findings", "procedures")

    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="D9E2F3")
    wrap = Alignment(wrap_text=True, vertical="top")

    for m in [str(x) for x in ws.merged_cells.ranges if start <= x.min_row <= stop]:
        ws.unmerge_cells(m)
    for r in range(start, stop):
        for c in range(2, 11):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).border = Border()

    PROC_END, FIND_START = 6, 7
    r = start
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=PROC_END)
    ws.merge_cells(start_row=r, start_column=FIND_START, end_row=r, end_column=10)
    w.put(ws, r, 2, "Procedures", "procedures")
    w.put(ws, r, FIND_START, "Findings", "procedures")
    for c in (2, FIND_START):
        ws.cell(row=r, column=c).font = Font(bold=True)
        ws.cell(row=r, column=c).fill = hdr_fill
    _boxrow(ws, r, box)
    r += 1

    for b, title, prows in blocks:
        if r >= stop - 1:
            break
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
        w.put(ws, r, 2, f"{b}. {title}", "procedures")
        ws.cell(row=r, column=2).font = Font(bold=True)
        ws.cell(row=r, column=2).fill = hdr_fill
        _boxrow(ws, r, box)
        r += 1
        for proc, finding in prows:
            if r >= stop - 1:
                break
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=PROC_END)
            ws.merge_cells(start_row=r, start_column=FIND_START, end_row=r, end_column=10)
            w.put(ws, r, 2, proc, "procedures")
            w.put(ws, r, FIND_START, finding, "procedures (llm)" if llm_used else "procedures")
            ws.cell(row=r, column=2).alignment = wrap
            ws.cell(row=r, column=FIND_START).alignment = wrap
            _boxrow(ws, r, box)
            ws.row_dimensions[r].height = max(15.0, 13.0 * (1 + max(len(proc), len(finding)) // 52))
            r += 1

    w.sheets_touched.add(ws.title)
    w.warnings.append(
        f"Audit Report: wrote a {len(blocks)}-block Procedures & Findings table "
        f"({r - start} rows){' - findings drafted by LLM' if llm_used else ''}."
    )


# --------------------------------------------------------------------------- #
# LLM annexure-mapping pass
# --------------------------------------------------------------------------- #
def _coerce(value):
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str):
        return None
    v = value.strip()
    if not v:
        return None
    if v.startswith("="):
        return v[:400]
    if _NUM_RE.match(v):
        neg = v.startswith("(") and v.endswith(")")
        try:
            n = float(v.strip("()%").replace(",", ""))
        except ValueError:
            return v[:200]
        return -n if neg else n
    return v[:200]


def _apply_llm_writes(w: _Writer, proposals: list[dict], done: set[tuple[str, str]]) -> None:
    applied = skipped = 0
    seen: set[tuple[str, str]] = set()
    for p in proposals:
        if "__error__" in p:
            w.warnings.append(f"LLM pass ({p.get('sheet', '?')}): {p['__error__']}")
            continue
        ws = w.sheet(str(p.get("sheet", "")))
        m = _CELL_RE.match(str(p.get("cell", "")).strip())
        if ws is None or not m:
            skipped += 1
            continue
        col = openpyxl.utils.column_index_from_string(m.group(1).upper())
        row = int(m.group(2))
        r, c = w._anchor(ws, row, col)
        ref = f"{get_column_letter(c)}{r}"
        if (ws.title, ref) in seen:
            continue
        seen.add((ws.title, ref))
        val = _coerce(p.get("value"))
        if val is None or r > ws.max_row + 5:
            skipped += 1
            continue

        cur = ws.cell(row=r, column=c).value
        # never touch an existing formula, and never overwrite a text
        # label/heading - only blank cells, numbers, or genuine ______ placeholders
        if isinstance(cur, str) and cur.startswith("="):
            skipped += 1
            continue
        if isinstance(cur, str) and cur.strip() and not _PLACEHOLDER_RE.search(cur):
            # allow the model to replace a short header-ish string only when it is
            # writing a formula there; otherwise protect the label
            if not (isinstance(val, str) and val.startswith("=") and len(cur) < 40):
                skipped += 1
                continue

        w.put(ws, r, c, val, f"llm:{llm_client.provider_name()}", inherit=True)
        applied += 1
    if applied or skipped:
        w.warnings.append(
            f"LLM annexure pass ({llm_client.provider_name()}): {applied} cell(s) written, {skipped} proposal(s) skipped."
        )


# --------------------------------------------------------------------------- #
# public entry point
# --------------------------------------------------------------------------- #
def generate_workbook(
    template_path: str | Path,
    out_path: str | Path,
    match: MatchResult,
    requirements: RequirementsResult | None,
    financials: FinancialData,
    project_name: str | None = None,
    period_end: str | None = None,
    context: EntityContext | None = None,
    preparer_explicit: bool = False,
    doc_paths: list | None = None,
) -> WorkbookReport:
    wb = openpyxl.load_workbook(str(template_path), data_only=False)
    w = _Writer(wb)
    ctx = context or EntityContext(jop_name=project_name or "", period_end=period_end or "")
    p = derive_period(requirements, ctx.jop_name or project_name, ctx.period_end or period_end)

    _apply_entity_and_period(w, p)
    _fill_pbc_checklist(w, match, requirements)
    _fill_requirements_tracker(w, match, requirements)
    # detail annexures first so the primary statements can reference them
    _fill_other_income(w, financials)
    _fill_bank_balances(w, financials)
    _fill_supplier_balances(w, financials)
    _fill_accruals(w, financials)
    _fill_income_statement(w, financials)
    _fill_balance_sheet(w, financials, p)
    _fill_project_details(w, ctx, p)
    _fill_signatory(w, ctx, explicit=preparer_explicit)
    _fill_notes_summary(w, match, financials)

    _llm_on = LLM_WORKBOOK and llm_client.available()
    try:
        _fill_audit_report_procedures(w, financials, ctx, match, use_llm=_llm_on)
    except Exception as exc:  # noqa: BLE001
        w.warnings.append(f"Audit Report procedures table skipped: {type(exc).__name__}: {exc}")

    # LLM pass: read every annexure's real layout and fill what the heuristics
    # missed, formulas and all - style-matched to the rows above.
    if LLM_WORKBOOK and llm_client.available():
        done_cells = {(cw.sheet, cw.cell) for cw in w.writes}
        try:
            _apply_llm_writes(
                w, workbook_llm.propose_writes(wb, ctx, financials, doc_paths or []), done_cells
            )
        except Exception as exc:  # noqa: BLE001
            w.warnings.append(f"LLM annexure pass skipped: {type(exc).__name__}: {exc}")

    for note in financials.notes:
        w.warnings.append(f"source: {note}")

    out_path = Path(out_path)
    wb.save(str(out_path))
    wb.close()

    return WorkbookReport(
        filename=out_path.name,
        sheets_touched=sorted(w.sheets_touched),
        writes=w.writes,
        formulas_written=w.formulas,
        warnings=w.warnings,
        context=ctx,
    )
