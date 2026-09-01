"""Session lifecycle + ingestion of the requirements sheet and master template."""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, UploadFile

from ..config import ALLOW_LOCAL_SAMPLES, SAMPLE_REQUIREMENTS, SAMPLE_TEMPLATE
from ..schemas import RequirementsResult, SessionInfo
from ..services.requirements_parser import parse_requirements
from ..store import store

router = APIRouter(prefix="/api", tags=["session"])


def _info(sess) -> SessionInfo:
    return SessionInfo(
        id=sess.id,
        has_requirements=sess.requirements is not None,
        has_template=sess.template_path is not None,
        document_count=len(sess.documents),
        has_match=sess.match is not None,
        has_workbook=sess.workbook_path is not None,
    )


def _save_upload(dest: Path, upload: UploadFile) -> None:
    dest.parent.mkdir(parents=True, exist_ok=True)
    with dest.open("wb") as fh:
        shutil.copyfileobj(upload.file, fh)


@router.post("/session", response_model=SessionInfo)
def create_session() -> SessionInfo:
    return _info(store.create())


@router.get("/session/{sid}", response_model=SessionInfo)
def get_session(sid: str) -> SessionInfo:
    try:
        return _info(store.get(sid))
    except KeyError:
        raise HTTPException(404, "session not found")


@router.delete("/session/{sid}")
def delete_session(sid: str) -> dict:
    store.delete(sid)
    return {"ok": True}


@router.get("/session/{sid}/requirements", response_model=RequirementsResult)
def get_requirements(sid: str) -> RequirementsResult:
    sess = _get(sid)
    if not sess.requirements:
        raise HTTPException(404, "no requirements uploaded")
    return sess.requirements


@router.get("/session/{sid}/template")
def get_template(sid: str) -> dict:
    sess = _get(sid)
    if not sess.template_path:
        raise HTTPException(404, "no template uploaded")
    return {"filename": sess.template_path.name, "sheets": sess.template_sheets}


@router.post("/session/{sid}/requirements", response_model=RequirementsResult)
def upload_requirements(
    sid: str,
    file: UploadFile = File(...),
    sheet: str | None = Form(default=None),
) -> RequirementsResult:
    sess = _get(sid)
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "requirements checklist must be an .xlsx file")
    dest = sess.dir / "requirements.xlsx"
    _save_upload(dest, file)
    result = _parse(dest, sheet, file.filename)
    sess.requirements = result
    sess.requirements_path = dest
    sess.match = None
    return result


@router.post("/session/{sid}/requirements/reparse", response_model=RequirementsResult)
def reparse_requirements(sid: str, sheet: str = Form(...)) -> RequirementsResult:
    """Re-read the already-uploaded checklist using a different worksheet."""
    sess = _get(sid)
    if not sess.requirements_path or not sess.requirements_path.exists():
        raise HTTPException(400, "no requirements checklist uploaded")
    name = sess.requirements.filename if sess.requirements else sess.requirements_path.name
    result = _parse(sess.requirements_path, sheet, name)
    sess.requirements = result
    sess.match = None
    return result


def _parse(dest, sheet, display_name):
    try:
        result = parse_requirements(dest, sheet=sheet)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(422, f"could not parse requirements sheet: {exc}")
    if not result.items:
        raise HTTPException(422, "no requirement line items detected in the sheet")
    result.filename = display_name or result.filename
    return result


@router.post("/session/{sid}/template")
def upload_template(sid: str, file: UploadFile = File(...)) -> dict:
    sess = _get(sid)
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "master template must be an .xlsx file")
    dest = sess.dir / "template.xlsx"
    _save_upload(dest, file)
    try:
        wb = openpyxl.load_workbook(dest, read_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(422, f"could not open template workbook: {exc}")
    sess.template_path = dest
    sess.template_sheets = sheets
    return {"filename": file.filename, "sheets": sheets}


@router.post("/session/{sid}/load-samples", response_model=SessionInfo)
def load_samples(sid: str) -> SessionInfo:
    """Developer convenience - ingest the bundled desktop sample files."""
    if not ALLOW_LOCAL_SAMPLES:
        raise HTTPException(403, "local samples are disabled")
    sess = _get(sid)
    req = Path(SAMPLE_REQUIREMENTS)
    tmpl = Path(SAMPLE_TEMPLATE)
    if req.exists():
        dest = sess.dir / "requirements.xlsx"
        shutil.copyfile(req, dest)
        parsed = parse_requirements(dest)
        parsed.filename = req.name
        sess.requirements = parsed
        sess.requirements_path = dest
    if tmpl.exists():
        dest = sess.dir / "template.xlsx"
        shutil.copyfile(tmpl, dest)
        wb = openpyxl.load_workbook(dest, read_only=True)
        sess.template_sheets = list(wb.sheetnames)
        wb.close()
        sess.template_path = dest
    return _info(sess)


def _get(sid: str):
    try:
        return store.get(sid)
    except KeyError:
        raise HTTPException(404, "session not found")
